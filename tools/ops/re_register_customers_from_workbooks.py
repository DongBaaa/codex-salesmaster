#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
FORM_DIR = ROOT / "양식"
DEFAULT_BASE_URL = "https://api.example.invalid"
TIMEOUT_SECONDS = 30

COLS = {
    "name": 0,
    "category": 1,
    "manager": 2,
    "staff": 3,
    "phone": 4,
    "mobile": 5,
    "fax": 6,
    "note": 7,
    "business_number": 8,
    "representative": 9,
    "business_type": 10,
    "business_item": 11,
    "branch_office": 12,
    "home_page": 13,
    "email": 14,
    "address1": 15,
    "address2": 16,
    "recipient": 17,
    "manage_number": 18,
    "total_receivable": 19,
}

TRADE_TYPE_MAP = {
    "": "매출",
    "매출": "매출",
    "판매": "매출",
    "매출처": "매출",
    "매입": "매입",
    "매입처": "매입",
    "매출/매입": "매출/매입",
    "판매/매입": "매출/매입",
    "매출매입": "매출/매입",
    "매입/매출처": "매출/매입",
}

NOTE_PREFIXES = {
    "manager": "담당자",
    "mobile": "휴대폰",
    "fax": "팩스번호",
    "branch_office": "종사업장",
    "home_page": "홈페이지",
    "recipient": "받는이",
    "manage_number": "관리번호",
    "total_receivable": "총미수금",
    "register_date": "등록일자",
}


@dataclass
class WorkbookSpec:
    alias: str
    tenant_code: str
    office_code: str
    filename: str
    username: str
    password: str


@dataclass
class WorkbookCustomerRow:
    alias: str
    tenant_code: str
    office_code: str
    name_original: str
    trade_type: str
    contact_person: str
    representative: str
    business_number: str
    business_type: str
    business_item: str
    address: str
    phone: str
    email: str
    notes: str
    branch_office: str
    location_signature: str
    exact_key: str
    fallback_keys: list[str]
    source_row_numbers: list[int]


@dataclass
class OfficeApplySummary:
    alias: str
    workbook_path: str
    workbook_source_rows: int
    workbook_merged_rows: int
    workbook_exact_duplicate_groups: int
    workbook_same_business_multi_name_groups: int
    server_rows_before: int
    server_rows_after: int
    created_count: int
    updated_count: int
    unchanged_count: int
    ambiguous_existing_count: int
    workbook_duplicate_merged_count: int
    verify_missing_count: int
    verify_extra_count: int
    workbook_business_multi_name_examples: list[dict[str, Any]]
    ambiguous_examples: list[dict[str, Any]]
    verify_missing_examples: list[dict[str, Any]]
    verify_extra_examples: list[dict[str, Any]]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return re.sub(r"[\s\-_./()\[\],]", "", text).upper()


def normalize_business_number(value: Any) -> str:
    return "".join(ch for ch in normalize_text(value) if ch.isalnum()).upper()


def normalize_trade_type(value: Any) -> str:
    normalized = normalize_text(value)
    return TRADE_TYPE_MAP.get(normalized, normalized or "매출")


def parse_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = normalize_text(value).replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal_text(value: Decimal) -> str:
    value = value.quantize(Decimal("1")) if value == value.to_integral() else value.normalize()
    if value == value.to_integral():
        return f"{int(value):,}"
    return format(value, "f").rstrip("0").rstrip(".")


def format_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    return text


def merge_unique_lines(*chunks: str) -> str:
    lines: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        if not chunk:
            continue
        for raw_line in chunk.splitlines():
            line = normalize_text(raw_line)
            if not line:
                continue
            key = line.casefold()
            if key in seen:
                continue
            seen.add(key)
            lines.append(line)
    return "\n".join(lines)


def combine_address(address1: Any, address2: Any) -> str:
    parts = [normalize_text(address1), normalize_text(address2)]
    return " ".join(part for part in parts if part).strip()


def build_location_signature(branch_office: Any, address: Any) -> str:
    return f"{normalize_key(branch_office)}|{normalize_key(address)}"


def build_exact_key(business_number: Any, name: Any, branch_office: Any, address: Any) -> str:
    biz = normalize_business_number(business_number)
    name_key = normalize_key(name)
    location = build_location_signature(branch_office, address)
    if biz and name_key and location != "|":
        return f"BIZ_NAME_LOC:{biz}|{name_key}|{location}"
    if not biz and name_key and location != "|":
        return f"NAME_LOC:{name_key}|{location}"
    return ""


def build_fallback_keys(business_number: Any, name: Any) -> list[str]:
    keys: list[str] = []
    biz = normalize_business_number(business_number)
    name_key = normalize_key(name)
    if biz and name_key:
        keys.append(f"BIZ_NAME:{biz}|{name_key}")
    if name_key:
        keys.append(f"NAME:{name_key}")
    if biz and not name_key:
        keys.append(f"BIZ:{biz}")
    return keys


def extract_note_value(notes: str, label: str) -> str:
    prefix = f"{label}:"
    for raw_line in notes.splitlines():
        line = normalize_text(raw_line)
        if line.startswith(prefix):
            return normalize_text(line[len(prefix):])
    return ""


def build_notes(row: list[Any]) -> str:
    lines: list[str] = []
    note_fields = [
        (NOTE_PREFIXES["manager"], row[COLS["manager"]]),
        (NOTE_PREFIXES["mobile"], row[COLS["mobile"]]),
        (NOTE_PREFIXES["fax"], row[COLS["fax"]]),
        (NOTE_PREFIXES["branch_office"], row[COLS["branch_office"]]),
        (NOTE_PREFIXES["home_page"], row[COLS["home_page"]]),
        (NOTE_PREFIXES["recipient"], row[COLS["recipient"]]),
        (NOTE_PREFIXES["manage_number"], row[COLS["manage_number"]]),
    ]
    for label, value in note_fields:
        text = normalize_text(value)
        if text:
            lines.append(f"{label}: {text}")

    total_receivable = parse_decimal(row[COLS["total_receivable"]])
    if total_receivable != 0:
        lines.append(f"{NOTE_PREFIXES['total_receivable']}: {format_decimal_text(total_receivable)}")

    note = normalize_text(row[COLS["note"]])
    if note:
        lines.append(note)

    return merge_unique_lines("\n".join(lines))


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def merge_workbook_row(base: WorkbookCustomerRow, incoming: WorkbookCustomerRow) -> WorkbookCustomerRow:
    merged = WorkbookCustomerRow(
        alias=base.alias,
        tenant_code=base.tenant_code,
        office_code=base.office_code,
        name_original=first_non_empty(base.name_original, incoming.name_original),
        trade_type=first_non_empty(base.trade_type, incoming.trade_type) or "매출",
        contact_person=first_non_empty(base.contact_person, incoming.contact_person),
        representative=first_non_empty(base.representative, incoming.representative),
        business_number=first_non_empty(base.business_number, incoming.business_number),
        business_type=first_non_empty(base.business_type, incoming.business_type),
        business_item=first_non_empty(base.business_item, incoming.business_item),
        address=first_non_empty(base.address, incoming.address),
        phone=first_non_empty(base.phone, incoming.phone),
        email=first_non_empty(base.email, incoming.email),
        notes=merge_unique_lines(base.notes, incoming.notes),
        branch_office=first_non_empty(base.branch_office, incoming.branch_office),
        location_signature=base.location_signature or incoming.location_signature,
        exact_key=base.exact_key or incoming.exact_key,
        fallback_keys=list(dict.fromkeys(base.fallback_keys + incoming.fallback_keys)),
        source_row_numbers=sorted(set(base.source_row_numbers + incoming.source_row_numbers)),
    )
    if not merged.location_signature:
        merged.location_signature = build_location_signature(merged.branch_office, merged.address)
    if not merged.exact_key:
        merged.exact_key = build_exact_key(merged.business_number, merged.name_original, merged.branch_office, merged.address)
    if not merged.fallback_keys:
        merged.fallback_keys = build_fallback_keys(merged.business_number, merged.name_original)
    return merged


def load_workbook_rows(spec: WorkbookSpec) -> tuple[list[WorkbookCustomerRow], dict[str, Any]]:
    workbook_path = FORM_DIR / spec.filename
    if not workbook_path.exists():
        raise FileNotFoundError(f"거래처 리스트 파일을 찾을 수 없습니다: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb.active
    raw_rows: list[WorkbookCustomerRow] = []
    source_row_count = 0

    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(values)
        if all(value is None or normalize_text(value) == "" for value in row):
            continue

        source_row_count += 1
        name = normalize_text(row[COLS["name"]])
        if not name:
            continue

        trade_type = normalize_trade_type(row[COLS["category"]])
        branch_office = normalize_text(row[COLS["branch_office"]])
        address = combine_address(row[COLS["address1"]], row[COLS["address2"]])
        business_number = normalize_text(row[COLS["business_number"]])
        notes = build_notes(row)
        exact_key = build_exact_key(business_number, name, branch_office, address)
        fallback_keys = build_fallback_keys(business_number, name)

        raw_rows.append(WorkbookCustomerRow(
            alias=spec.alias,
            tenant_code=spec.tenant_code,
            office_code=spec.office_code,
            name_original=name,
            trade_type=trade_type,
            contact_person=normalize_text(row[COLS["staff"]]),
            representative=normalize_text(row[COLS["representative"]]),
            business_number=business_number,
            business_type=normalize_text(row[COLS["business_type"]]),
            business_item=normalize_text(row[COLS["business_item"]]),
            address=address,
            phone=normalize_text(row[COLS["phone"]]),
            email=normalize_text(row[COLS["email"]]),
            notes=notes,
            branch_office=branch_office,
            location_signature=build_location_signature(branch_office, address),
            exact_key=exact_key,
            fallback_keys=fallback_keys,
            source_row_numbers=[row_number],
        ))

    grouped_rows: dict[str, WorkbookCustomerRow] = {}
    duplicate_groups = 0
    duplicate_rows_merged = 0
    for row in raw_rows:
        group_key = row.exact_key or f"FALLBACK|{normalize_business_number(row.business_number)}|{normalize_key(row.name_original)}|{row.location_signature}"
        if group_key not in grouped_rows:
            grouped_rows[group_key] = row
            continue
        duplicate_groups += 1
        duplicate_rows_merged += 1
        grouped_rows[group_key] = merge_workbook_row(grouped_rows[group_key], row)

    merged_rows = list(grouped_rows.values())
    business_groups: dict[str, set[str]] = defaultdict(set)
    for row in merged_rows:
        biz = normalize_business_number(row.business_number)
        if biz:
            business_groups[biz].add(row.name_original)

    multi_name_groups = [
        {"business_number": business_number, "names": sorted(names)}
        for business_number, names in business_groups.items()
        if len(names) > 1
    ]
    multi_name_groups.sort(key=lambda current: (-len(current["names"]), current["business_number"]))

    audit = {
        "workbook_path": str(workbook_path),
        "source_row_count": source_row_count,
        "merged_row_count": len(merged_rows),
        "duplicate_groups": duplicate_groups,
        "duplicate_rows_merged": duplicate_rows_merged,
        "business_multi_name_group_count": len(multi_name_groups),
        "business_multi_name_examples": multi_name_groups[:10],
    }
    return merged_rows, audit


class ApiClient:
    def __init__(self, base_url: str, username: str, password: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.token = ""

    def login(self) -> None:
        response = self.session.post(
            self.base_url + "/auth/login",
            json={"username": self.username, "password": self.password},
            timeout=TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        payload = response.json()
        self.token = payload.get("accessToken") or payload.get("AccessToken") or ""
        if not self.token:
            raise RuntimeError(f"{self.username} 로그인 응답에 accessToken이 없습니다.")
        self.session.headers.update({"Authorization": f"Bearer {self.token}"})

    def get_customers(self) -> list[dict[str, Any]]:
        response = self.session.get(self.base_url + "/customers", params={"take": 5000}, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, list):
            raise RuntimeError(f"{self.username} 거래처 조회 응답 형식이 예상과 다릅니다.")
        return payload

    def create_customer(self, payload: dict[str, Any]) -> dict[str, Any]:
        response = self.session.post(self.base_url + "/customers", json=payload, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
        return response.json()

    def update_customer(self, customer_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        response = self.session.put(self.base_url + f"/customers/{customer_id}", json=payload, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
        return response.json()


def customer_exact_key(customer: dict[str, Any]) -> str:
    branch_office = extract_note_value(normalize_text(customer.get("notes")), NOTE_PREFIXES["branch_office"])
    return build_exact_key(
        customer.get("businessNumber"),
        customer.get("nameOriginal"),
        branch_office,
        customer.get("address"),
    )


def customer_fallback_keys(customer: dict[str, Any]) -> list[str]:
    return build_fallback_keys(customer.get("businessNumber"), customer.get("nameOriginal"))


def register_customer_lookup(lookup: dict[str, dict[str, Any] | None], customer: dict[str, Any]) -> None:
    exact_key = customer_exact_key(customer)
    if exact_key:
        lookup[exact_key] = customer

    for key in dict.fromkeys(customer_fallback_keys(customer)):
        if key not in lookup:
            lookup[key] = customer
            continue

        existing = lookup.get(key)
        if existing is not None and normalize_text(existing.get("id")) != normalize_text(customer.get("id")):
            lookup[key] = None


def build_customer_lookup(customers: list[dict[str, Any]]) -> dict[str, dict[str, Any] | None]:
    lookup: dict[str, dict[str, Any] | None] = {}
    for customer in customers:
        register_customer_lookup(lookup, customer)
    return lookup


def find_matching_customer(row: WorkbookCustomerRow, lookup: dict[str, dict[str, Any] | None]) -> tuple[dict[str, Any] | None, str]:
    if row.exact_key:
        candidate = lookup.get(row.exact_key)
        if candidate is not None:
            return candidate, "exact"
        if row.exact_key in lookup:
            return None, "exact_ambiguous"

    for key in row.fallback_keys:
        candidate = lookup.get(key)
        if candidate is not None:
            return candidate, f"fallback:{key.split(':', 1)[0]}"
        if key in lookup:
            return None, f"ambiguous:{key.split(':', 1)[0]}"

    return None, "new"


def trim_optional(value: Any) -> str:
    return normalize_text(value)


def build_name_match_key(name: str) -> str:
    return trim_optional(name).upper()


def build_customer_payload(row: WorkbookCustomerRow, existing: dict[str, Any] | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "tenantCode": row.tenant_code,
        "officeCode": row.office_code,
        "responsibleOfficeCode": row.office_code,
        "nameOriginal": row.name_original,
        "nameMatchKey": build_name_match_key(row.name_original),
        "tradeType": row.trade_type,
        "department": trim_optional(existing.get("department") if existing else ""),
        "contactPerson": row.contact_person,
        "representative": row.representative,
        "businessNumber": row.business_number,
        "businessType": row.business_type,
        "businessItem": row.business_item,
        "address": row.address,
        "phone": row.phone,
        "email": row.email,
        "notes": row.notes,
        "isDeleted": False,
    }

    if existing:
        payload["id"] = existing.get("id")
        payload["revision"] = existing.get("revision", 0)
        payload["customerMasterId"] = existing.get("customerMasterId")
        payload["categoryId"] = existing.get("categoryId")
    return payload


def normalize_payload_for_compare(payload: dict[str, Any]) -> dict[str, str]:
    return {
        "tenantCode": trim_optional(payload.get("tenantCode")),
        "officeCode": trim_optional(payload.get("officeCode")),
        "responsibleOfficeCode": trim_optional(payload.get("responsibleOfficeCode")),
        "nameOriginal": trim_optional(payload.get("nameOriginal")),
        "nameMatchKey": trim_optional(payload.get("nameMatchKey")),
        "tradeType": trim_optional(payload.get("tradeType")),
        "department": trim_optional(payload.get("department")),
        "contactPerson": trim_optional(payload.get("contactPerson")),
        "representative": trim_optional(payload.get("representative")),
        "businessNumber": trim_optional(payload.get("businessNumber")),
        "businessType": trim_optional(payload.get("businessType")),
        "businessItem": trim_optional(payload.get("businessItem")),
        "address": trim_optional(payload.get("address")),
        "phone": trim_optional(payload.get("phone")),
        "email": trim_optional(payload.get("email")),
        "notes": trim_optional(payload.get("notes")),
        "categoryId": trim_optional(payload.get("categoryId")),
        "customerMasterId": trim_optional(payload.get("customerMasterId")),
    }


def select_office_customers(customers: list[dict[str, Any]], office_code: str) -> list[dict[str, Any]]:
    normalized_office = office_code.upper()
    return [
        customer for customer in customers
        if trim_optional(customer.get("responsibleOfficeCode")).upper() == normalized_office
        and not bool(customer.get("isDeleted"))
    ]


def build_existing_exact_duplicate_count(customers: list[dict[str, Any]]) -> int:
    counter = Counter(customer_exact_key(customer) for customer in customers if customer_exact_key(customer))
    return sum(1 for count in counter.values() if count > 1)


def sample_rows(rows: list[WorkbookCustomerRow], limit: int = 10) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows[:limit]:
        result.append({
            "name_original": row.name_original,
            "business_number": row.business_number,
            "address": row.address,
            "branch_office": row.branch_office,
            "source_row_numbers": row.source_row_numbers,
        })
    return result


def sample_customers(customers: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for customer in customers[:limit]:
        result.append({
            "id": customer.get("id"),
            "name_original": customer.get("nameOriginal"),
            "business_number": customer.get("businessNumber"),
            "address": customer.get("address"),
            "responsible_office_code": customer.get("responsibleOfficeCode"),
        })
    return result


def compare_workbook_and_server(rows: list[WorkbookCustomerRow], customers: list[dict[str, Any]]) -> tuple[list[WorkbookCustomerRow], list[dict[str, Any]]]:
    workbook_keys = {row.exact_key for row in rows if row.exact_key}
    server_by_key: dict[str, dict[str, Any]] = {}
    for customer in customers:
        exact_key = customer_exact_key(customer)
        if exact_key and exact_key not in server_by_key:
            server_by_key[exact_key] = customer

    missing_rows = [row for row in rows if row.exact_key and row.exact_key not in server_by_key]
    extra_customers = [customer for customer in customers if customer_exact_key(customer) and customer_exact_key(customer) not in workbook_keys]
    return missing_rows, extra_customers


def apply_office(spec: WorkbookSpec, rows: list[WorkbookCustomerRow], base_url: str, apply_changes: bool) -> tuple[OfficeApplySummary, dict[str, Any]]:
    client = ApiClient(base_url, spec.username, spec.password)
    client.login()
    office_customers_before = select_office_customers(client.get_customers(), spec.office_code)
    lookup = build_customer_lookup(office_customers_before)

    created_count = 0
    updated_count = 0
    unchanged_count = 0
    ambiguous_count = 0
    ambiguous_examples: list[dict[str, Any]] = []
    action_log: list[dict[str, Any]] = []

    for row in rows:
        matched, match_mode = find_matching_customer(row, lookup)
        payload = build_customer_payload(row, matched)

        if matched is None and match_mode.startswith("ambiguous"):
            ambiguous_count += 1
            example = {
                "name_original": row.name_original,
                "business_number": row.business_number,
                "address": row.address,
                "branch_office": row.branch_office,
                "match_mode": match_mode,
                "source_row_numbers": row.source_row_numbers,
            }
            if len(ambiguous_examples) < 10:
                ambiguous_examples.append(example)
            action_log.append({"action": "create", **example})
            created_count += 1
            if not apply_changes:
                continue
            created = client.create_customer(payload)
            office_customers_before.append(created)
            register_customer_lookup(lookup, created)
            continue

        if matched is None:
            action_log.append({
                "action": "create",
                "name_original": row.name_original,
                "business_number": row.business_number,
                "address": row.address,
                "branch_office": row.branch_office,
                "source_row_numbers": row.source_row_numbers,
            })
            created_count += 1
            if not apply_changes:
                continue
            created = client.create_customer(payload)
            office_customers_before.append(created)
            register_customer_lookup(lookup, created)
            continue

        current_compare = normalize_payload_for_compare(build_customer_payload(row, matched))
        existing_compare = normalize_payload_for_compare(matched)
        if current_compare == existing_compare:
            unchanged_count += 1
            continue

        action_log.append({
            "action": "update",
            "match_mode": match_mode,
            "customer_id": matched.get("id"),
            "name_original": row.name_original,
            "business_number": row.business_number,
            "address": row.address,
            "branch_office": row.branch_office,
            "source_row_numbers": row.source_row_numbers,
        })
        updated_count += 1
        if not apply_changes:
            continue
        updated = client.update_customer(trim_optional(matched.get("id")), payload)
        for index, customer in enumerate(office_customers_before):
            if trim_optional(customer.get("id")) == trim_optional(updated.get("id")):
                office_customers_before[index] = updated
                break
        register_customer_lookup(lookup, updated)

    office_customers_after = select_office_customers(client.get_customers(), spec.office_code)
    missing_rows, extra_customers = compare_workbook_and_server(rows, office_customers_after)

    business_groups: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        biz = normalize_business_number(row.business_number)
        if biz:
            business_groups[biz].add(row.name_original)
    multi_name_examples = [
        {"business_number": business_number, "names": sorted(names)}
        for business_number, names in business_groups.items()
        if len(names) > 1
    ]
    multi_name_examples.sort(key=lambda current: (-len(current["names"]), current["business_number"]))

    summary = OfficeApplySummary(
        alias=spec.alias,
        workbook_path=str(FORM_DIR / spec.filename),
        workbook_source_rows=0,
        workbook_merged_rows=len(rows),
        workbook_exact_duplicate_groups=0,
        workbook_same_business_multi_name_groups=len(multi_name_examples),
        server_rows_before=len(office_customers_before) if not apply_changes else 0,
        server_rows_after=len(office_customers_after),
        created_count=created_count,
        updated_count=updated_count,
        unchanged_count=unchanged_count,
        ambiguous_existing_count=ambiguous_count,
        workbook_duplicate_merged_count=0,
        verify_missing_count=len(missing_rows),
        verify_extra_count=len(extra_customers),
        workbook_business_multi_name_examples=multi_name_examples[:10],
        ambiguous_examples=ambiguous_examples,
        verify_missing_examples=sample_rows(missing_rows),
        verify_extra_examples=sample_customers(extra_customers),
    )

    debug = {
        "office_customers_before": sample_customers(select_office_customers(client.get_customers(), spec.office_code), 20) if not apply_changes else [],
        "office_customers_after": sample_customers(office_customers_after, 20),
        "existing_exact_duplicate_count": build_existing_exact_duplicate_count(office_customers_after),
        "action_log": action_log[:100],
    }
    return summary, debug


def hydrate_summary(summary: OfficeApplySummary, audit: dict[str, Any], server_rows_before: int) -> OfficeApplySummary:
    summary.workbook_source_rows = int(audit["source_row_count"])
    summary.workbook_exact_duplicate_groups = int(audit["duplicate_groups"])
    summary.workbook_duplicate_merged_count = int(audit["duplicate_rows_merged"])
    summary.workbook_same_business_multi_name_groups = int(audit["business_multi_name_group_count"])
    summary.workbook_business_multi_name_examples = list(audit["business_multi_name_examples"])
    summary.server_rows_before = server_rows_before
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="아이티월드/유즈넷 거래처 리스트를 라이브 서버에 재등록합니다.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--mode", choices=["dry-run", "apply"], default="dry-run")
    parser.add_argument("--output", default="")
    parser.add_argument("--itworld-username", default="itworld")
    parser.add_argument("--itworld-password", default="1234")
    parser.add_argument("--usenet-username", default="usenet")
    parser.add_argument("--usenet-password", default="1234")
    return parser.parse_args()


def build_default_output_path(mode: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return ROOT / "artifacts" / "reports" / f"customer-workbook-reregister-{mode}-{timestamp}.json"


def main() -> int:
    args = parse_args()
    output_path = Path(args.output) if args.output else build_default_output_path(args.mode)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    specs = [
        WorkbookSpec(
            alias="ITWORLD",
            tenant_code="ITWORLD",
            office_code="ITWORLD",
            filename="아이티월드 거래처 리스트.xlsx",
            username=args.itworld_username,
            password=args.itworld_password,
        ),
        WorkbookSpec(
            alias="USENET",
            tenant_code="USENET_GROUP",
            office_code="USENET",
            filename="유즈넷 거래처 리스트.xlsx",
            username=args.usenet_username,
            password=args.usenet_password,
        ),
    ]

    report: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "mode": args.mode,
        "base_url": args.base_url,
        "offices": {},
    }

    for spec in specs:
        rows, audit = load_workbook_rows(spec)

        audit_client = ApiClient(args.base_url, spec.username, spec.password)
        audit_client.login()
        server_rows_before = len(select_office_customers(audit_client.get_customers(), spec.office_code))

        summary, debug = apply_office(spec, rows, args.base_url, apply_changes=args.mode == "apply")
        summary = hydrate_summary(summary, audit, server_rows_before)
        report["offices"][spec.alias] = {
            "summary": asdict(summary),
            "audit": audit,
            "debug": debug,
        }

    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8-sig")
    print(str(output_path))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except requests.HTTPError as exc:
        response = exc.response
        detail = ""
        if response is not None:
            try:
                detail = response.text
            except Exception:
                detail = ""
        print(f"HTTP 오류: {exc}\n{detail}", file=sys.stderr)
        raise
    except Exception as exc:
        print(f"실패: {exc}", file=sys.stderr)
        raise
